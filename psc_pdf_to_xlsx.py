#!/usr/bin/env python3
"""
Kerala PSC Ranked List PDF → XLSX Converter
============================================

Converts Kerala Public Service Commission ranked-list PDFs (e.g. Assistant
Salesman, LDC, etc.) into a clean, structured Excel workbook with:
    - One row per candidate
    - A Category column (Main List / each Supplementary community / each
      Differently Abled category) since rank numbers repeat across lists
    - A Summary sheet with metadata and per-category counts

Usage
-----
    python psc_pdf_to_xlsx.py <input.pdf> [output.xlsx]

If output.xlsx is omitted, the script writes "<input>_converted.xlsx" next
to the input file.

Requirements
------------
    pip install pdfplumber openpyxl

How it works
------------
1. pdfplumber extracts every word with its (x, y) position.
2. Words are grouped into visual lines by y-coordinate.
3. The "Rank Reg.No Name DOB Commy Remarks" header line is detected on each
   page; its word positions define the column x-boundaries (so the script
   adapts if the PSC ever tweaks column widths).
4. Each subsequent line is classified:
     * starts with a 1-3 digit rank AND has a 7-digit reg-no  → data row
     * matches a known section title (Main List, Ezhava/Thiyya/Billava, …)
                                                              → category switch
     * otherwise, if the previous row had a wrapped cell      → continuation
     * boilerplate (NOTE, page numbers, footer, etc.)         → skipped
5. Words on a data row are routed to columns purely by x-position, so a name
   like "MARY VINEETHA SHAJU" stays in the Name column even if the community
   field is blank.
6. The result is written to XLSX with formatting, banded categories,
   freeze-pane, auto-filter, and a Summary sheet.
"""

from __future__ import annotations

import re
import sys
from collections import OrderedDict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Section headers found in PSC ranked lists. Keys are matched case-insensitively
# against the full line text. Value = category label written to the XLSX (or
# None for headers that are just dividers and should not produce a category).
# ---------------------------------------------------------------------------
SECTION_HEADERS: "OrderedDict[str, str | None]" = OrderedDict([
    ("main list", "Main List"),
    ("supplementary list", None),                      # divider only
    ("list of differently abled candidates", None),    # divider only
    ("ezhava/thiyya/billava", "Supplementary - Ezhava/Thiyya/Billava"),
    ("scheduled caste converts to christianity",
     "Supplementary - Scheduled Caste Converts to Christianity"),
    ("scheduled caste", "Supplementary - Scheduled Caste"),
    ("scheduled tribe", "Supplementary - Scheduled Tribe"),
    ("muslim", "Supplementary - Muslim"),
    ("latin catholics/a.i", "Supplementary - Latin Catholics/A.I"),
    ("obc", "Supplementary - OBC"),
    ("viswakarma", "Supplementary - Viswakarma"),
    ("siuc nadar", "Supplementary - SIUC Nadar"),
    ("dheevara", "Supplementary - Dheevara"),
    ("hindu nadar", "Supplementary - Hindu Nadar"),
    ("economically weaker section", "Supplementary - Economically Weaker Section"),
    ("low vision", "Differently Abled - Low Vision"),
    ("hard of hearing", "Differently Abled - Hard of Hearing"),
    ("ld including cp", "Differently Abled - LD/CP/LC/Dw/AAV"),
    ("asd (m), sld, mi", "Differently Abled - ASD/SLD/MI/Multiple"),
])

# Lines starting with these tokens are boilerplate and should be ignored.
SKIP_LINE_PREFIXES = (
    "note", "page", "ranked list", "number of candidates",
    "total number of candidates", "kerala public", "office of",
    "(approved", "section officer", "by order", "george k",
    "district officer", "e - ezhava", "rank reg.no", "no.",
)

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
REGNO_RE = re.compile(r"^\d{7}$")
RANK_RE = re.compile(r"^\d{1,3}$")


# ---------------------------------------------------------------------------
# PDF text helpers
# ---------------------------------------------------------------------------
def group_words_by_line(words: list[dict], y_tol: float = 3.0) -> list[list[dict]]:
    """Cluster words into visual lines using their `top` y-coordinate."""
    if not words:
        return []
    words = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))
    lines: list[list[dict]] = []
    current: list[dict] = [words[0]]
    anchor = words[0]["top"]
    for w in words[1:]:
        if abs(w["top"] - anchor) <= y_tol:
            current.append(w)
        else:
            lines.append(sorted(current, key=lambda x: x["x0"]))
            current = [w]
            anchor = w["top"]
    lines.append(sorted(current, key=lambda x: x["x0"]))
    return lines


def detect_column_zones(header_line: list[dict]) -> list[tuple[str, float, float]] | None:
    """
    Build column (name, x_left, x_right) zones from the header row.

    The PSC PDF center-aligns the header text over its column but left-aligns
    the data underneath, so simply using each header's x0 as a boundary makes
    data drift one column to the left (e.g. dates at x≈383 fall under a "DOB"
    header that starts at x≈396 and get pulled into the Name column).

    We therefore place the boundary between two columns at the midpoint of the
    gap between adjacent header words. This is robust to small layout changes
    and to the header being center-aligned over a wide column.
    """
    found: dict[str, dict] = {}
    for w in header_line:
        t = w["text"].lower().rstrip(".")
        if t == "rank":                       found["rank"] = w
        elif t in ("reg.no", "regno", "reg"): found["regno"] = w
        elif t == "name":                     found["name"] = w
        elif t == "dob":                      found["dob"] = w
        elif t in ("commy", "community"):     found["community"] = w
        elif t == "remarks":                  found["remarks"] = w

    order = ["rank", "regno", "name", "dob", "community", "remarks"]
    present = [k for k in order if k in found]
    if not {"rank", "regno", "name", "dob", "community"}.issubset(present):
        return None

    zones: list[tuple[str, float, float]] = []
    for i, name in enumerate(present):
        w = found[name]
        # Column boundary is placed 2/3 of the way from the LEFT header's right
        # edge to the RIGHT header's left edge.  This works because:
        #   - data is left-aligned, headers are center-aligned, so a column's
        #     data can start well to the left of its header's x0;
        #   - biasing the boundary toward the right header gives each column
        #     extra room on its right side to absorb wrap-continuation text;
        #   - it still leaves a buffer so left-aligned data of the right column
        #     stays in the right column.
        if i == 0:
            left = 0.0
        else:
            prev = found[present[i - 1]]
            left = (2 * prev["x1"] + w["x0"]) / 3
        if i == len(present) - 1:
            right = 1e6
        else:
            nxt = found[present[i + 1]]
            right = (2 * w["x1"] + nxt["x0"]) / 3
        zones.append((name, left, right))
    return zones


def assign_word_to_zone(word: dict, zones: list[tuple[str, float, float]]) -> str:
    """Return the column whose x-zone contains the word's x0 (with tolerance)."""
    x = word["x0"]
    for name, left, right in zones:
        if left - 2 <= x < right:
            return name
    return zones[-1][0]   # fall through → rightmost column


def parse_row(line: list[dict],
              zones: list[tuple[str, float, float]]) -> dict[str, str]:
    """Distribute the words on a line into rank/regno/name/dob/community/remarks."""
    buckets: dict[str, list[str]] = {k: [] for k in
                                     ("rank", "regno", "name", "dob",
                                      "community", "remarks")}
    for w in line:
        buckets[assign_word_to_zone(w, zones)].append(w["text"])
    return {k: " ".join(v).strip() for k, v in buckets.items()}


def line_text(line: list[dict]) -> str:
    return " ".join(w["text"] for w in line).strip()


def _smart_join(prev: str, more: str) -> str:
    """
    Join two fragments of a wrapped cell.

    Drop the joining space when the wrap clearly happened mid-token:
      * `prev` ends with '-'  (e.g. "OBC-" + "VEERASAIVAS")
      * `prev` has an unclosed '['  (e.g. "VEERASAIVAS[YOG" + "ISYOGEESWARA]")

    Otherwise join with a single space.
    """
    if not prev:
        return more
    if not more:
        return prev
    inside_bracket = prev.count("[") > prev.count("]")
    if prev.endswith("-") or inside_bracket:
        return prev + more
    return prev + " " + more


def match_section(text: str) -> tuple[bool, str | None]:
    """
    Return (is_section_header, category_or_None).
    The boolean indicates we should *consume* the line; the category is what
    to set on subsequent rows (None means: it was a divider, keep prior cat).
    """
    t = text.lower().strip()
    for key, label in SECTION_HEADERS.items():
        if t == key or t.startswith(key):
            return True, label
    return False, None


def is_boilerplate(text: str) -> bool:
    t = text.lower().strip()
    if not t:
        return True
    if t.isdigit() and len(t) <= 3:
        return True   # standalone page number
    return any(t.startswith(p) for p in SKIP_LINE_PREFIXES)


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------
def extract_rank_list(
    pdf_path: Path,
    progress_callback=None,
) -> tuple[list[dict], dict[str, str]]:
    """
    Extract every candidate from a PSC ranked-list PDF.

    Parameters
    ----------
    pdf_path
        Path to the input PDF.
    progress_callback
        Optional ``callable(current_page, total_pages, status_message)`` called
        once per page so a UI can drive a progress bar. ``status_message`` is a
        short human-readable label like ``"Reading page 7 of 23"``.

    Returns
    -------
    (rows, meta)
        ``rows`` is a list of dicts (one per candidate) with the seven output
        columns; ``meta`` is a dict of harvested document metadata.
    """
    rows: list[dict] = []
    meta: dict[str, str] = {}
    current_category: str | None = None
    zones: list[tuple[str, float, float]] | None = None
    pending: dict[str, str] | None = None

    def flush():
        nonlocal pending
        if pending is not None:
            rows.append(pending)
            pending = None

    with pdfplumber.open(str(pdf_path)) as pdf:
        total_pages = len(pdf.pages)
        for page_idx, page in enumerate(pdf.pages, start=1):
            if progress_callback is not None:
                progress_callback(
                    page_idx, total_pages,
                    f"Reading page {page_idx} of {total_pages}",
                )

            words = page.extract_words(use_text_flow=False,
                                       keep_blank_chars=False)
            lines = group_words_by_line(words)

            for line in lines:
                text = line_text(line)
                low = text.lower()

                # ---- harvest a bit of metadata from page 1 -----------------
                if not meta:
                    if "ranked list no." in low:
                        m = re.search(r"ranked list no\.?\s*:?\s*(\S+)", low)
                        if m: meta["ranked_list_no"] = m.group(1).upper()
                    if "cat. no." in low or "cat.no." in low:
                        m = re.search(r"cat\.?\s*no\.?\s*:?\s*(\S+)", low)
                        if m: meta["category_no"] = m.group(1)
                    if "omr test" in low and "held on" in low:
                        m = re.search(r"held on\s+(\d{2}\.\d{2}\.\d{4})", low)
                        if m: meta["omr_test_date"] = m.group(1)
                    if "brought into force" in low:
                        m = re.search(r"with effect from\s+(\d{2}\.\d{2}\.\d{4})", low)
                        if m: meta["in_force_from"] = m.group(1)

                # ---- detect the table header --------------------------------
                if "rank" in low and "reg.no" in low and "name" in low:
                    found = detect_column_zones(line)
                    if found:
                        zones = found
                    flush()
                    continue

                if zones is None:
                    continue   # haven't seen the header on this page yet

                # ---- skip boilerplate ---------------------------------------
                if is_boilerplate(text):
                    flush()
                    continue

                # ---- section header? ----------------------------------------
                is_sec, label = match_section(text)
                if is_sec:
                    flush()
                    if label is not None:
                        current_category = label
                    continue

                # ---- data row? ----------------------------------------------
                parsed = parse_row(line, zones)

                # pdfplumber occasionally glues a date and the community text
                # into one token, e.g. "16/07/2001OBC-VILAKKITHALA". Split it.
                glued = re.match(r"^(\d{2}/\d{2}/\d{4})(\S.*)$", parsed["dob"])
                if glued and not parsed["community"]:
                    parsed["dob"] = glued.group(1)
                    parsed["community"] = glued.group(2)

                # Sanity-check the DOB column: it must contain exactly one date.
                # Any non-date tokens are misclassified neighbours. Tokens that
                # appear before the date belong to Name (long names like
                # "NEENU STENSLAVOUS K S" can push the trailing initial past
                # the Name→DOB boundary). Tokens after the date belong to
                # Community.
                date_in_dob = re.search(r"(\d{2}/\d{2}/\d{4})", parsed["dob"])
                if date_in_dob:
                    before = parsed["dob"][:date_in_dob.start()].strip()
                    after  = parsed["dob"][date_in_dob.end():].strip()
                    parsed["dob"] = date_in_dob.group(1)
                    if before:
                        parsed["name"] = (parsed["name"] + " " + before).strip()
                    if after:
                        parsed["community"] = (after + " " + parsed["community"]).strip()

                rank_ok = bool(RANK_RE.match(parsed["rank"]))
                regno_ok = bool(REGNO_RE.match(parsed["regno"].replace(" ", "")))
                if rank_ok and regno_ok:
                    flush()
                    pending = {
                        "Category": current_category or "Uncategorised",
                        "Rank":      int(parsed["rank"]),
                        "Reg.No":    parsed["regno"].replace(" ", ""),
                        "Name":      parsed["name"],
                        "DOB":       parsed["dob"],
                        "Community": parsed["community"],
                        "Remarks":   parsed["remarks"],
                    }
                    continue

                # ---- continuation of previous row (wrapped cell) ------------
                # Continuation lines only ever extend the Name or Community
                # fields, never Remarks. Anything that landed in Remarks here
                # is actually wrapped Community text that drifted right.
                if pending is not None:
                    cont = parse_row(line, zones)
                    wrapped_community = " ".join(
                        v for v in (cont["community"], cont["remarks"]) if v
                    ).strip()
                    if cont["name"]:
                        pending["Name"] = _smart_join(pending["Name"], cont["name"])
                    if wrapped_community:
                        pending["Community"] = _smart_join(
                            pending["Community"], wrapped_community
                        )
            # end of page
            flush()

    # tidy: collapse double spaces in text fields
    for r in rows:
        for k in ("Name", "Community", "Remarks"):
            r[k] = re.sub(r"\s+", " ", r[k]).strip()

    return rows, meta


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------
def write_xlsx(rows: list[dict], meta: dict[str, str], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked List"

    # ---- Title block --------------------------------------------------------
    ws["A1"] = "KERALA PSC RANKED LIST"
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    info_bits = []
    if meta.get("ranked_list_no"):  info_bits.append(f"Ranked List No.: {meta['ranked_list_no']}")
    if meta.get("category_no"):     info_bits.append(f"Cat. No.: {meta['category_no']}")
    if meta.get("omr_test_date"):   info_bits.append(f"OMR: {meta['omr_test_date']}")
    if meta.get("in_force_from"):   info_bits.append(f"In force from: {meta['in_force_from']}")
    ws["A2"] = "  |  ".join(info_bits) if info_bits else ""
    ws.merge_cells("A2:G2")
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", start_color="2E75B6")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ---- Header row ---------------------------------------------------------
    headers = ["Category", "Rank", "Reg.No", "Name", "DOB", "Community", "Remarks"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HEADER_ROW].height = 26

    # ---- Body ---------------------------------------------------------------
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_cat = None
    band = False
    for i, row in enumerate(rows, start=HEADER_ROW + 1):
        if row["Category"] != current_cat:
            current_cat = row["Category"]
            band = not band
        fill_color = "F2F2F2" if band else "FFFFFF"
        values = [row[h] for h in headers]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=i, column=col_idx, value=v)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.fill = PatternFill("solid", start_color=fill_color)
            if col_idx in (2, 3, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left",
                                        vertical="center", wrap_text=True)

    widths = {"A": 42, "B": 7, "C": 11, "D": 32, "E": 12, "F": 28, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = f"B{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:G{HEADER_ROW + len(rows)}"

    # ---- Summary sheet ------------------------------------------------------
    s = wb.create_sheet("Summary")
    s["A1"] = "Summary"
    s.merge_cells("A1:B1")
    s["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    s["A1"].fill = PatternFill("solid", start_color="1F4E79")
    s["A1"].alignment = Alignment(horizontal="center")
    s.row_dimensions[1].height = 22

    info = [
        ("Ranked List No.", meta.get("ranked_list_no", "")),
        ("Category No.",    meta.get("category_no", "")),
        ("OMR Test Date",   meta.get("omr_test_date", "")),
        ("In force from",   meta.get("in_force_from", "")),
        ("", ""),
        ("List Section", "Count"),
    ]
    r = 2
    for label, value in info:
        a = s.cell(row=r, column=1, value=label)
        b = s.cell(row=r, column=2, value=value)
        a.font = Font(name="Arial", bold=True, size=11)
        b.font = Font(name="Arial", size=11)
        if label == "List Section":
            for cell in (a, b):
                cell.fill = PatternFill("solid", start_color="2E75B6")
                cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        r += 1

    counts: "OrderedDict[str, int]" = OrderedDict()
    for row in rows:
        counts[row["Category"]] = counts.get(row["Category"], 0) + 1
    start_count_row = r
    for cat, cnt in counts.items():
        s.cell(row=r, column=1, value=cat).font = Font(name="Arial", size=10)
        c = s.cell(row=r, column=2, value=cnt)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        r += 1

    s.cell(row=r, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    tot = s.cell(row=r, column=2, value=f"=SUM(B{start_count_row}:B{r - 1})")
    tot.font = Font(name="Arial", bold=True, size=11)
    tot.alignment = Alignment(horizontal="center")
    for col in ("A", "B"):
        s[f"{col}{r}"].fill = PatternFill("solid", start_color="FFE699")
    s.column_dimensions["A"].width = 55
    s.column_dimensions["B"].width = 14

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 1
    pdf_path = Path(argv[1]).expanduser().resolve()
    if not pdf_path.is_file():
        print(f"Error: {pdf_path} not found", file=sys.stderr)
        return 2
    out_path = (Path(argv[2]).expanduser().resolve()
                if len(argv) > 2
                else pdf_path.with_name(pdf_path.stem + "_converted.xlsx"))

    print(f"Reading: {pdf_path}")
    rows, meta = extract_rank_list(pdf_path)
    print(f"Extracted {len(rows)} candidates across "
          f"{len({r['Category'] for r in rows})} sections")
    write_xlsx(rows, meta, out_path)
    print(f"Wrote:   {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
